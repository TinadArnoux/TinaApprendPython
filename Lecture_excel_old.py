"""import pandas as pd

# Lire le fichier Excel en spécifiant les types de données
df = pd.read_excel('chargementPosts.xlsx', dtype={'texte': str, 'date': 'datetime64[ns]'})

# Afficher les données de la colonne "texte"
print(df['texte'])

# Filtrer les données pour les dates après une certaine date
date_seuil = pd.to_datetime('2023-01-01')
df_filtre = df[df['date'] > date_seuil]
print(df_filtre)"""

import datetime
import os

import openpyxl

# magestion de BDD
from gestionBDD import Database

# Chemin vers votre fichier .db
db = Database('Mes_Tweets.db')
print("BDD : Mes_Tweets.db")


class FichierExcel:
    def deplacer_et_renommer_fichier(self, ancien_chemin, nouveau_chemin, nouveau_nom):
        # déplace et renomme le fichier
        # Construire le nouveau chemin complet
        nouveau_chemin_complet = os.path.join(nouveau_chemin, nouveau_nom)
        print("new nom " + nouveau_nom)
        print("new chem " + nouveau_chemin)
        print("anc che " + ancien_chemin)
        print("new complet " + nouveau_chemin_complet)
        # Renommer le fichier (ce qui le déplace)
        try:
            os.rename(ancien_chemin, nouveau_chemin_complet)
            print(f"Le fichier a été déplacé et renommé en {nouveau_chemin_complet}")
        except OSError as error:
            print(f"Une erreur s'est produite : {error}")

    def ouvre_et_lit_fichier(self, fichier):
        try:
            workbook = openpyxl.load_workbook(fichier)
            # Sélectionner la feuille
            sheet = workbook.active

            # Itérer sur les lignes
            for row in sheet.iter_rows(values_only=True):
                texte, date = row
                print(texte, date)
#            return [texte, date]  # return à revoir pour donner un tableau
        except Exception as e:
            print("Une erreur s'est produite : " + str(e))


    def cree_fichier_vide(self):
        # créer un fichier vide
        workbook = openpyxl.Workbook()
        # Enregistrer le classeur
        workbook.save(fichierComplet)


# Ouvrir le fichier Excel
fichier = 'chargementPosts'
fichierComplet = fichier + ".xlsx"
# formater les nouveau nom et chemin
# Obtenir la date actuelle
date_aujourdhui = (datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S"))
ancien_chemin = "C:/Users/mchri/PycharmProjects/PythonProject1/" + fichierComplet
nouveau_chemin = "C:/Users/mchri/PycharmProjects/PythonProject1/archive_excel/"
nouveau_nom = f"{fichier}_traite_{date_aujourdhui}.xlsx"

# lire le fichier
ouvre_et_lit_fichier(fichierComplet)
# déplacer le fichier lu
deplacer_et_renommer_fichier(ancien_chemin, nouveau_chemin, nouveau_nom)
