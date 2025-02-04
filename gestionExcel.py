import os
import openpyxl

class FichierExcel:
    @staticmethod
    def deplacer_et_renommer_fichier(ancien_chemin, nouveau_chemin, nouveau_nom):
        # déplace et renomme le fichier
        # Construire le nouveau chemin complet
        nouveau_chemin_complet = os.path.join(nouveau_chemin, nouveau_nom)
        print("new nom " + nouveau_nom)
        print("new chem " + nouveau_chemin)
        print("anc che " + ancien_chemin)
        print("new complet " + nouveau_chemin_complet)
        # Renommer le fichier (ce qui le déplace)
        try:
            os.rename(ancien_chemin, nouveau_chemin_complet)
            print(f"Le fichier a été déplacé et renommé en {nouveau_chemin_complet}")
        except OSError as error:
            print(f"Une erreur s'est produite : {error}")

    @staticmethod
    def ouvre_et_lit_fichier(fichier):
        try:
            workbook = openpyxl.load_workbook(fichier)
            # Sélectionner la feuille
            sheet = workbook.active

            # Itérer sur les lignes
            for row in sheet.iter_rows(values_only=True):
                texte, date = row
                print(texte, date)
#            return [texte, date]  # return à revoir pour donner un tableau
        except Exception as e:
            print("Une erreur s'est produite : " + str(e))

    @staticmethod
    def cree_fichier_vide(fichiercomplet):
        # créer un fichier vide
        workbook = openpyxl.Workbook()
        # Enregistrer le classeur
        workbook.save(fichiercomplet)